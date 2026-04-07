"""
ETF Holdings Scraper
Downloads holdings data from ETF providers (iShares, SSGA/SPDR).
These are the index/ETF funds that UC invests in where holdings are public.

Usage:
    python Finance/scrapers/etf_holdings.py

Output:
    Finance/outputs/etf_<fund_name>.csv  (one per fund)
"""

import csv
import io
import json
import os
import sys

import requests

try:
    import openpyxl
except ImportError:
    openpyxl = None

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCES_PATH = os.path.join(BASE_DIR, "sources.json")
SOURCES_DIR = os.path.join(BASE_DIR, "sources")
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
}


def download_file(url, filename):
    """Download file to sources directory."""
    path = os.path.join(SOURCES_DIR, filename)
    if os.path.exists(path):
        print(f"  Already downloaded: {filename}")
        return path
    print(f"  Downloading: {filename}")
    resp = requests.get(url, headers=HEADERS, timeout=60)
    resp.raise_for_status()
    with open(path, "wb") as f:
        f.write(resp.content)
    return path


def parse_ishares_csv(file_path):
    """
    Parse iShares holdings CSV.
    iShares CSVs have metadata rows at the top before the actual data.
    """
    holdings = []

    with open(file_path, "r", encoding="utf-8-sig") as f:
        lines = f.readlines()

    # Find the header row (contains "Ticker" and "Name")
    header_idx = None
    for i, line in enumerate(lines):
        if "Ticker" in line and "Name" in line and "Weight" in line:
            header_idx = i
            break

    if header_idx is None:
        print(f"  Warning: Could not find header row in {file_path}")
        return []

    # Parse from header row onwards, skipping blank/incomplete rows
    reader = csv.DictReader(lines[header_idx:])
    for row in reader:
        if row is None:
            continue
        name = row.get("Name")
        if name is None or not name.strip() or name.strip() == "-":
            continue

        name = name.strip()
        ticker = (row.get("Ticker") or "").strip()
        weight_str = (row.get("Weight (%)") or row.get("Weight") or "").strip()
        market_value = (row.get("Market Value") or "").strip()

        weight_val = 0.0
        if weight_str:
            try:
                weight_val = float(weight_str.replace(",", "").replace("%", ""))
            except ValueError:
                weight_val = 0.0

        # iShares weights are already in percent, convert to decimal
        holdings.append({
            "name": name,
            "ticker": ticker,
            "weight": weight_val / 100.0,
            "market_value": market_value,
        })

    return holdings


def parse_ssga_xlsx(file_path):
    """
    Parse SSGA/SPDR holdings XLSX file.
    SSGA files have metadata rows before the data table.
    """
    if openpyxl is None:
        print("  Warning: openpyxl not installed, skipping XLSX parsing")
        print("  Install with: pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    holdings = []

    # Find header row
    header_idx = None
    headers = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_vals = [str(c).strip() if c else "" for c in row]
        if "Name" in row_vals and ("Weight" in row_vals or "Ticker" in row_vals):
            headers = {v: j for j, v in enumerate(row_vals) if v}
            header_idx = i
            break

    if header_idx is None:
        print(f"  Warning: Could not find header in {file_path}")
        return []

    # Parse data rows
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue

        row_vals = [c for c in row]
        name_idx = headers.get("Name", None)
        if name_idx is None or name_idx >= len(row_vals):
            continue

        name = str(row_vals[name_idx]).strip() if row_vals[name_idx] else ""
        if not name or name == "None":
            continue

        ticker = ""
        weight = 0.0

        if "Ticker" in headers and headers["Ticker"] < len(row_vals):
            ticker = str(row_vals[headers["Ticker"]] or "").strip()

        for weight_key in ["Weight", "Weight (%)", "% Of Fund"]:
            if weight_key in headers and headers[weight_key] < len(row_vals):
                try:
                    w = row_vals[headers[weight_key]]
                    weight = float(str(w).replace(",", "").replace("%", ""))
                    if weight > 1:
                        weight /= 100.0
                except (ValueError, TypeError):
                    pass
                break

        holdings.append({
            "name": name,
            "ticker": ticker,
            "weight": weight,
        })

    wb.close()
    return holdings


def write_etf_csv(holdings, output_path, fund_name):
    """Write ETF holdings to a standardized CSV."""
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["name", "ticker", "weight"])
        writer.writeheader()
        # Sort by weight descending
        sorted_h = sorted(holdings, key=lambda x: x.get("weight", 0), reverse=True)
        for h in sorted_h:
            writer.writerow({
                "name": h["name"],
                "ticker": h.get("ticker", ""),
                "weight": h.get("weight", 0),
            })

    print(f"  {fund_name}: {len(holdings)} holdings -> {output_path}")


def main():
    os.makedirs(SOURCES_DIR, exist_ok=True)
    os.makedirs(OUTPUTS_DIR, exist_ok=True)

    with open(SOURCES_PATH) as f:
        sources = json.load(f)

    print("=== ETF Holdings Scraper ===\n")

    etf_sources = sources.get("etf_holdings", {})

    for fund_key, url in etf_sources.items():
        print(f"\nProcessing: {fund_key}")

        ext = "xlsx" if url.endswith(".xlsx") else "csv"
        source_file = download_file(url, f"etf_{fund_key}.{ext}")

        if ext == "xlsx":
            holdings = parse_ssga_xlsx(source_file)
        else:
            holdings = parse_ishares_csv(source_file)

        if holdings:
            output_path = os.path.join(OUTPUTS_DIR, f"etf_{fund_key}.csv")
            write_etf_csv(holdings, output_path, fund_key)
        else:
            print(f"  Warning: No holdings parsed for {fund_key}")

    print("\nDone!")


if __name__ == "__main__":
    main()
